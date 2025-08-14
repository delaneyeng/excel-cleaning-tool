
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy
from datetime import datetime
import tempfile
import os

def is_red_bold(cell):
    if not cell.font or not cell.font.bold:
        return False
    color = cell.font.color
    if not color or not hasattr(color, 'rgb') or not color.rgb:
        return False
    return str(cell.font.color.rgb).upper().endswith("FF0000")

def is_green_font(cell):
    return (
        cell.font and
        cell.font.color and
        hasattr(cell.font.color, 'rgb') and
        cell.font.color.rgb and
        str(cell.font.color.rgb).upper().endswith("00B050")
    )

def clean_workbook(file):
    wb = openpyxl.load_workbook(file)
    if "Data List" not in wb.sheetnames:
        st.error("'Data List' worksheet not found.")
        return None

    ws = wb["Data List"]

    for cell_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(cell_range))
    ws.delete_rows(2)
    ws.delete_rows(2)
    ws["A1"] = "S/N"

    for col in range(ws.max_column, 1, -1):
        ws.cell(row=1, column=col + 1).value = ws.cell(row=1, column=col).value
        ws.cell(row=1, column=col).value = None

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = 'd/m/yyyy h:mm AM/PM '

    ws.insert_cols(3)
    for row in range(1, ws.max_row + 1):
        for attr in ['font', 'border', 'fill', 'number_format', 'protection', 'alignment']:
            setattr(ws.cell(row=row, column=3), attr, copy(getattr(ws.cell(row=row, column=2), attr)))
    ws["B1"] = "Date/Time"
    ws["C1"] = "Days"

    day_markers = {
        290: "Day 1",
        578: "Day 2",
        866: "Day 3",
        1154: "Day 4",
        1442: "Day 5",
        1730: "Day 6",
        2018: "Day 7"
    }
    for r, label in day_markers.items():
        ws.cell(row=r, column=3).value = label

    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=4).value = ws.cell(row=row, column=2).value
    ws.delete_cols(4)

    latest_time = None
    latest_cell = None
    for col in range(2, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        try:
            dt = cell.value if isinstance(cell.value, datetime) else datetime.strptime(str(cell.value).strip(), "%d/%m/%Y %I:%M %p")
            if not latest_time or dt > latest_time:
                latest_time = dt
                latest_cell = cell
        except:
            continue

    red_bold = Font(bold=True, color="FF0000")
    if latest_cell:
        latest_cell.font = red_bold
        ws.cell(row=2, column=latest_cell.column + 1).font = red_bold
        for row in range(2, ws.max_row + 1):
            for col in range(2, ws.max_column):
                cell = ws.cell(row=row, column=col)
                try:
                    val = cell.value if isinstance(cell.value, datetime) else datetime.strptime(str(cell.value).strip(), "%d/%m/%Y %I:%M %p")
                    if val == latest_time:
                        cell.font = red_bold
                        ws.cell(row=row, column=col + 1).font = red_bold
                except:
                    continue

    green_font = Font(name='Calibri', size=11, bold=False, italic=False, color="00B050")
    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column):
            cell = ws.cell(row=row, column=col)
            if is_red_bold(cell):
                for r in range(2, row):
                    top_cell = ws.cell(row=r, column=col)
                    right_cell = ws.cell(row=r, column=col + 1)
                    try:
                        val = top_cell.value
                        if isinstance(val, datetime):
                            top_cell.font = green_font
                            right_cell.font = green_font
                        else:
                            for fmt in ("%d/%m/%Y %I:%M %p", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
                                try:
                                    datetime.strptime(str(val).strip(), fmt)
                                    top_cell.font = green_font
                                    right_cell.font = green_font
                                    break
                                except:
                                    continue
                    except:
                        continue

    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column):
            cell = ws.cell(row=row, column=col)
            if is_green_font(cell) and not is_red_bold(cell):
                cell.value = None

    for col in range(2, ws.max_column):
        if col == 3:
            continue
        values = []
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                values.append((cell.value, copy(cell.font), copy(cell.fill), copy(cell.number_format)))
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).font = Font()
            ws.cell(row=row, column=col).fill = PatternFill()
            ws.cell(row=row, column=col).number_format = 'General'
        for i, (val, font, fill, num_fmt) in enumerate(values):
            cell = ws.cell(row=2 + i, column=col)
            cell.value = val
            cell.font = font
            cell.fill = fill
            cell.number_format = num_fmt

    for col_idx in [16, 14, 12, 10, 8, 6, 4]:
        if col_idx <= ws.max_column:
            ws.delete_cols(col_idx)

    reset_font = Font(name="Segoe UI", size=9, bold=False, color="000000")
    for col in range(1, ws.max_column + 1):
        ws.cell(row=2, column=col).font = reset_font

    for i, row in enumerate(range(2, ws.max_row + 1), start=1):
        ws.cell(row=row, column=1).value = i

    return wb

st.title("🧹 Excel Cleaning Automation Tool")

uploaded_file = st.file_uploader("Upload your Excel file (.xlsx only)", type=["xlsx"])

if uploaded_file is not None:
    with st.spinner("Processing your file..."):
        cleaned_wb = clean_workbook(uploaded_file)
        if cleaned_wb:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                cleaned_wb.save(tmp.name)
                tmp.seek(0)
                st.success("✅ File cleaned successfully!")
                st.download_button(
                    label="📥 Download Cleaned Excel File",
                    data=tmp.read(),
                    file_name="Cleaned_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
